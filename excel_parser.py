import openpyxl
from collections import defaultdict
import pandas as pd

class Parser:
    
    def parse_tree(self, sheet, filepath):

        #filepath = "POSCO 2PL Mill Control System_20221106_NEW-21_MJW01_STN1.xlsx"

        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        # Define a dictionary to store the parsed data for each node
        node_data_name = []
        node_data_dict = {}


        """
        Excel 시트의 트리 구조를 파싱합니다.
        """
        max_row = sheet.max_row
        max_col = sheet.max_column
    
        def find_end_row(start_row, col):
            """
            주어진 시작 행과 열을 기준으로 해당 열의 끝 행을 찾습니다.
            """
            end_row = start_row
            while end_row <= max_row and (sheet.cell(row=end_row, column=col).value is None or "@End" not in str(sheet.cell(row=end_row, column=col).value)):
                end_row += 1
            return end_row-1

        def find_end_data_row(start_row, col):
            """
            주어진 시작 행과 열을 기준으로 해당 열의 끝 행을 찾습니다.
            """
            end_row = start_row
            while end_row <= max_row and (sheet.cell(row=end_row, column=col).value is not None  and "@" not in str(sheet.cell(row=end_row, column=col).value)):
                end_row += 1
            return end_row-1

        def parse_node(row, col):
            """
            노드와 해당 자식들을 재귀적으로 파싱합니다.
            """
            if row > max_row or col > max_col:
                return None

            #노드의 이름을 저장.
            node_name = sheet.cell(row=row, column=col).value
            if not node_name or node_name.startswith('@End'):
                return None
            
            
            # 현재 노드의 끝 행 정의
            end_row = find_end_row(row, col)
            
            # 노드의 데이터 파싱
            current_search_row = row + 1
            col_start = col + 1
            col_end = col_start
            data = defaultdict(list)
            end_data_row = end_row

            while current_search_row < max_row and sheet.cell(current_search_row, col_end).value is not None:
                # 컬럼 이름에서 "(*" , "*)" 문자를 제거 후 저장.
                column_name = sheet.cell(current_search_row, col_end).value
                if column_name.startswith('(*'):
                    column_name = column_name[2:]
                if column_name.endswith('*)'):
                    column_name = column_name[:-2]
                column_name = column_name.strip()

                
                #첫번째 컬럼에서 데이터의 끝 열을 찾음
                if(col_end-col == 1):
                    end_data_row = find_end_data_row(current_search_row,col_end)
                    #print(f"node name : {node_name} end_data : {end_data_row}")
                
                current_data_row = current_search_row + 1
                while current_data_row <= end_data_row:
                    
                    data[column_name].append(sheet.cell(current_data_row, col_end).value)
                    current_data_row += 1


                col_end += 1

            df = pd.DataFrame(data)
            node_data_name.append(node_name)
            node_data_dict[node_name] = df
            
            #동일한 트리 레벨의 데이터 파싱
            if current_search_row < max_row and sheet.cell(end_data_row+2, col).value is not None and "@" in str(sheet.cell(end_data_row+2, col).value):
                parse_node(end_data_row+2, col)
            #하위 트리 레벨의 데이터 파싱
            if current_search_row < max_row and sheet.cell(end_data_row+1, col+1).value is not None and "@" in str(sheet.cell(end_data_row+1, col+1).value):
                parse_node(end_data_row+1, col+1)  
            # 파싱한 최종 데이터를 pandas DataFrame으로 변환하고 출력
        root = parse_node(1, 1)

        return node_data_name, node_data_dict


# parse_tree(sheet)

# for i in node_data_name:
#     print(node_data_dict[i])
